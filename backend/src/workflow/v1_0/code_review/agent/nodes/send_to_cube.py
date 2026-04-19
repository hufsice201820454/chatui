"""노드 6: Excel 생성 → CDN 업로드(Mock) → Cube 채널 전송(Mock)

Mock 처리 항목:
  - CDN 업로드 : 로컬 output/ 디렉터리에 파일 저장 후 가상 CDN URL 반환
  - Cube 전송  : 사내 Cube API 호출을 시뮬레이션 (실제 HTTP 호출 없음)
"""
import logging
import os
import re
from datetime import datetime
from typing import Optional
from uuid import uuid4

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from agent.state import AgentState

logger = logging.getLogger(__name__)

_OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "output",
)
_CDN_BASE_URL = "https://cdn.internal/code-review"


# ── Excel 생성 ────────────────────────────────────────────────────────────────

def _parse_markdown_table(md: str) -> tuple[list[str], list[list[str]]]:
    """마크다운 테이블을 헤더와 행 목록으로 파싱합니다."""
    lines = [l.strip() for l in md.strip().splitlines() if l.strip().startswith("|")]
    if len(lines) < 3:
        return [], []

    def _split_row(line: str) -> list[str]:
        return [cell.strip() for cell in line.strip("|").split("|")]

    headers = _split_row(lines[0])
    # lines[1]은 구분선 (---|---) → 건너뜀
    rows = [_split_row(line) for line in lines[2:] if not re.match(r"^\|[-| ]+\|$", line)]
    return headers, rows


def _create_excel(state: AgentState) -> Optional[str]:
    """분석 결과를 Excel 파일로 생성하고 로컬 경로를 반환합니다."""
    os.makedirs(_OUTPUT_DIR, exist_ok=True)

    headers, rows = _parse_markdown_table(state.get("final_answer", ""))
    if not headers:
        logger.warning("[Node6] 마크다운 테이블 파싱 실패 — 빈 Excel을 생성합니다.")
        headers = ["Project", "Source File", "Line", "Issue Detail", "Reason", "Recommended Solution"]
        rows = []

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    project_key = state.get("project_key", "unknown").replace("/", "_")
    filename = f"code_review_{project_key}_{timestamp}.xlsx"
    filepath = os.path.join(_OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Code Review"

    # ── 헤더 스타일 ──────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # ── 데이터 행 ────────────────────────────────────────────────────────────
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for row_idx, row in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # ── 열 너비 자동 조정 ────────────────────────────────────────────────────
    col_widths = {1: 20, 2: 30, 3: 8, 4: 40, 5: 45, 6: 50}
    for col_idx, width in col_widths.items():
        if col_idx <= len(headers):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info("[Node6] Excel 생성 완료: %s (%d행)", filepath, len(rows))
    return filepath


# ── CDN 업로드 (Mock) ─────────────────────────────────────────────────────────

def _mock_cdn_upload(filepath: str) -> str:
    """
    [Mock] CDN에 파일을 업로드합니다.
    실제 업로드 없이 가상 CDN URL을 반환합니다.
    """
    filename = os.path.basename(filepath)
    cdn_url = f"{_CDN_BASE_URL}/{filename}"

    logger.info("[Mock CDN] 업로드 시뮬레이션 — 파일: %s", filename)
    logger.info("[Mock CDN] CDN URL: %s", cdn_url)

    # 실제 구현 시 아래와 같은 형태로 교체:
    # import boto3
    # s3 = boto3.client("s3")
    # s3.upload_file(filepath, BUCKET_NAME, filename)
    # cdn_url = f"https://{CDN_DOMAIN}/{filename}"

    return cdn_url


# ── Cube 채널 전송 (Mock) ─────────────────────────────────────────────────────

def _mock_cube_send(channel: str, cdn_url: str, project_key: str, analysis_date: str) -> dict:
    """
    [Mock] Cube 채널로 파일 링크를 전송합니다.
    실제 Cube API 호출 없이 전송 성공을 시뮬레이션합니다.
    """
    message_id = f"msg_{uuid4().hex[:10]}"
    payload = {
        "channel": channel,
        "message": (
            f"[Code Review 분석 결과]\n"
            f"프로젝트: {project_key}\n"
            f"분석일시: {analysis_date}\n"
            f"파일 다운로드: {cdn_url}"
        ),
        "file_url": cdn_url,
        "message_type": "file_share",
    }

    logger.info("[Mock Cube] 전송 시뮬레이션 — 채널: %s", channel)
    logger.info("[Mock Cube] Payload: %s", payload)

    # 실제 구현 시 아래와 같은 형태로 교체:
    # import requests
    # response = requests.post(
    #     f"{CUBE_API_URL}/channels/{channel}/messages",
    #     headers={"Authorization": f"Bearer {CUBE_TOKEN}"},
    #     json=payload,
    # )
    # response.raise_for_status()

    return {
        "success": True,
        "channel": channel,
        "message_id": message_id,
        "cdn_url": cdn_url,
        "sent_at": datetime.now().isoformat(),
    }


# ── 노드 함수 ─────────────────────────────────────────────────────────────────

def send_to_cube(state: AgentState) -> AgentState:
    """
    분석 결과를 Excel로 생성 → CDN 업로드 → Cube 채널로 전송합니다.
    """
    channel = state.get("cube_channel", "")
    project_key = state.get("project_key", "unknown")
    analysis_date = state.get("analysis_date", "")

    # ── Step 1: Excel 생성 ────────────────────────────────────────────────────
    try:
        excel_path = _create_excel(state)
    except Exception as e:
        logger.error("[Node6] Excel 생성 실패: %s", e)
        return {
            **state,
            "cube_send_result": {"success": False, "error": f"Excel 생성 실패: {e}"},
        }

    # ── Step 2: CDN 업로드 (Mock) ─────────────────────────────────────────────
    try:
        cdn_url = _mock_cdn_upload(excel_path)
    except Exception as e:
        logger.error("[Node6] CDN 업로드 실패: %s", e)
        return {
            **state,
            "excel_file_path": excel_path,
            "cube_send_result": {"success": False, "error": f"CDN 업로드 실패: {e}"},
        }

    # ── Step 3: Cube 채널 전송 (Mock) ─────────────────────────────────────────
    try:
        result = _mock_cube_send(channel, cdn_url, project_key, analysis_date)
    except Exception as e:
        logger.error("[Node6] Cube 전송 실패: %s", e)
        return {
            **state,
            "excel_file_path": excel_path,
            "cdn_url": cdn_url,
            "cube_send_result": {"success": False, "error": f"Cube 전송 실패: {e}"},
        }

    logger.info("[Node6] Cube 채널 전송 완료 — 채널: %s, 메시지ID: %s", channel, result["message_id"])
    return {
        **state,
        "excel_file_path": excel_path,
        "cdn_url": cdn_url,
        "cube_send_result": result,
    }
